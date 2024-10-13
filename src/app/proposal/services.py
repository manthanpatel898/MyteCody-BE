import asyncio
import base64
import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import partial
import json
from multiprocessing import Process
import os
import smtplib
import threading
from bson import ObjectId
from src.app.healper.response import make_response
from src.app.proposal.schema import SaveStep1Prompt
from src.app.utils.constants import PROPOSAL_STATUS
from src.app.utils.openai_functions import generate_text, generate_text_json, generate_text_json_with_4o, generate_text_with_4o_for_conversation
from src.app.utils.pdf import design_pdf, sanitize_filename
from src.db import db
import concurrent.futures
import json as std_json
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from src.app.utils.prompts import (
    step1InitialPrompt,
    step1SystemContext,
    step1AssistantContext,
    step2AssistantContext,
    step2InitialPrompt,
    step2SystemContext,
    step3AssistantContext,
    step3InitialPrompt,
    step3SystemContext,
    step4AssistantContext,
    step4InitialPrompt,
    step4SystemContext,
    conversation_assistant_context,
    conversation_initial_prompt,
    conversation_system_context,
    define_epics_system_context,
    define_epics_assistant_context,
    define_epics_initial_prompt,
    user_stories_assistant_context,
    user_stories_initial_prompt,
    user_stories_system_context,
    task_definition_assistant_context,
    task_definition_initial_prompt,
    task_definition_system_context,
    task_complexity_assistant_context,
    task_complexity_initial_prompt,
    task_complexity_options,
    task_complexity_system_context,
    proposal_assistant_context,
    proposal_initial_prompt,
    proposal_json_format,
    proposal_system_context,
    milestone_assistant_context,
    milestone_initial_prompt,
    milestone_json_format,
    milestone_system_context,
    risks_assistant_context,
    risks_initial_prompt,
    risks_json_format,
    risks_system_context,
    titleInitialPromot,
    titleAssistantContext,
    titleSystemContext,
    descriptionInitialPromot,
    descriptionAssistantContext,
    descriptionSystemContext
)

from src.app.utils.messages import (
    BUSINESS_VERTICAL_FETCH_SUCCESS,
    BUSINESS_VERTICAL_GENERATION_SUCCESS,
    BUSINESS_VERTICAL_NOT_FOUND,
    BUSINESS_VERTICAL_UPDATE_SUCCESS,
    CLIENT_ONBOARDING_COMPLETED,
    CLIENT_ONBOARDING_IN_PROGRESS,
    CONVERSATION_NOT_FOUND,
    CONVERSATION_RETRIEVED_SUCCESSFULLY,
    CONVERSATION_SAVED_SUCCESSFULLY,
    CONVERSATION_SUCCESS,
    EPIC_ADDED_SUCCESS,
    EPIC_DELETED_SUCCESS,
    EPIC_NOT_FOUND,
    EPIC_UPDATED_SUCCESS,
    EPICS_FETCHED_SUCCESS,
    EPICS_GENERATED,
    EPICS_GENERATION_FAILED,
    EPICS_GENERATION_STARTED,
    EPICS_GENERATION_SUCCESS,
    GENERATING_STORIES,
    GENERATING_TASKS,
    NO_CONVERSATION_FOUND,
    PROJECT_DATA_MISSING,
    PROJECT_VISION_GENERATION_SUCCESS,
    PROJECT_VISION_NOT_FOUND,
    PROJECT_VISION_SUCCESS,
    PROPOSAL_NOT_FOUND,
    INTERNAL_SERVER_ERROR,
    REVENUE_MODEL_FETCH_SUCCESS,
    REVENUE_MODEL_GENERATION_SUCCESS,
    REVENUE_MODEL_NOT_FOUND,
    SAVE_PROJECT_VISION_SUCCESS,
    STAKEHOLDER_EPIC_OR_STORY_NOT_FOUND,
    STAKEHOLDER_GENERATION_SUCCESS,
    STAKEHOLDER_NOT_FOUND,
    STAKEHOLDER_OR_EPIC_NOT_FOUND,
    STAKEHOLDERS_FETCH_SUCCESS,
    STAKEHOLDERS_NOT_FOUND,
    STAKEHOLDERS_UPDATED_SUCCESS,
    STORIES_FETCHED_SUCCESS,
    STORIES_GENERATED_SUCCESS,
    STORIES_GENERATION_FAILED,
    STORY_ADDED_SUCCESS,
    STORY_DELETED_SUCCESS,
    STORY_NOT_FOUND,
    STORY_UPDATED_SUCCESS,
    TASK_ADDED_SUCCESS,
    TASK_DELETED_SUCCESS,
    TASK_NOT_FOUND,
    TASK_UPDATED_SUCCESS,
    TASKS_FETCHED_SUCCESS,
    TASKS_GENERATED_SUCCESS,
    USER_NOT_FOUND
)

def chat_conversation(payload, user_id):
    """GENERATE CHAT CONVERSATION RESPONSES"""
    try:
        # Check if global_generated_id exists, if not, create it with sr_no 0
        global_id = db.global_generated_id.find_one({})
        if not global_id:
            db.global_generated_id.insert_one({"sr_no": 0})
            global_id = {"sr_no": 0}

        if not payload.get("proposal_id"):
            # First-time conversation
            thread = [
                {
                    "SalesRep": "Hey, glad you're here with us. Can you start by telling us what you plan on building as a platform?"
                },
                {"UserResponse": payload["data"]},
            ]

            sr_no = global_id["sr_no"]
            db.global_generated_id.update_one({}, {"$inc": {"sr_no": 1}})
            created_at = updated_at = datetime.datetime.utcnow()

            # Insert a new proposal with created_at and updated_at
            new_proposal = {
                "user": user_id,
                "sr_no": sr_no + 1,
                "status": str(PROPOSAL_STATUS["in_progress"]),
                "step": 0,
                "last_step": CLIENT_ONBOARDING_IN_PROGRESS,
                "created_at": created_at,
                "updated_at": updated_at
            }
            result = db.proposals.insert_one(new_proposal)
            proposal_id = result.inserted_id

            # Generate input message for title and description
            title_input_message = titleInitialPromot.replace("{conversation_thread}", payload["data"])
            description_input_message = descriptionInitialPromot.replace("{conversation_thread}", str(thread))

            # Generate title and description
            title = generate_text(titleSystemContext, titleAssistantContext, title_input_message, user_id, proposal_id)
            description = generate_text(descriptionSystemContext, descriptionAssistantContext, description_input_message, user_id, proposal_id)

            # Update the proposal with the generated title, description, and updated_at
            updated_at = datetime.datetime.utcnow()

            db.proposals.find_one_and_update(
                {"_id": proposal_id, "user": user_id},
                {
                    "$set": {
                        "title": title,
                        "description": description,
                        "updated_at": updated_at
                    }
                }
            )

            # Insert new conversation with created_at and updated_at
            db.conversation.insert_one(
                {
                    "data": thread,
                    "user_id": user_id,
                    "proposal_id": proposal_id,
                    "created_at": created_at,
                    "updated_at": updated_at
                }
            )

        else:
            # Continuing conversation
            proposal_id = ObjectId(payload["proposal_id"])
            updated_at = datetime.datetime.utcnow()

            db.conversation.find_one_and_update(
                {"user_id": user_id, "proposal_id": proposal_id},
                {
                    "$push": {"data": {"UserResponse": payload["data"]}},
                    "$set": {"updated_at": updated_at}
                }
            )

            conversation = db.conversation.find_one({"user_id": user_id, "proposal_id": proposal_id})
            if not conversation:
                return make_response(
                    status="error",
                    message=PROPOSAL_NOT_FOUND,
                    status_code=404
                )

            thread = conversation["data"]
            thread.append({"UserResponse": payload["data"]})

            # Generate input message for title and description
            title_input_message = titleInitialPromot.replace("{conversation_thread}", str(thread))
            description_input_message = descriptionInitialPromot.replace("{conversation_thread}", str(thread))

            # Generate title and description
            title = generate_text(titleSystemContext, titleAssistantContext, title_input_message, user_id, proposal_id)
            description = generate_text(descriptionSystemContext, descriptionAssistantContext, description_input_message, user_id, proposal_id)

            # Update the proposal with the generated title, description, and updated_at
            updated_at = datetime.datetime.utcnow()

            db.proposals.find_one_and_update(
                {"_id": proposal_id, "user": user_id},
                {
                    "$set": {
                        "title": title,
                        "description": description,
                        "updated_at": updated_at
                    }
                }
            )

        # Generate next conversation response
        input_message = conversation_initial_prompt.replace("{conversation_thread}", str(thread))
        prompt = generate_text_with_4o_for_conversation(
            conversation_system_context,
            conversation_assistant_context,
            input_message,
            user_id,
            proposal_id
        )

        db.conversation.find_one_and_update(
            {"user_id": user_id, "proposal_id": proposal_id},
            {
                "$push": {"data": {"SalesRep": prompt}},
                "$set": {"updated_at": updated_at}
            }
        )

        return make_response(
            status="success",
            message=CONVERSATION_SUCCESS,
            data={"proposal_id": str(proposal_id), "data": prompt},
            status_code=200
        )
    except Exception as e:
        print(f"Error in client onboarding: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            status_code=500
        )

def get_conversation(user_id, proposal_id):
    """
    Service to retrieve user conversation for a specific proposal.

    Parameters:
        user_id (str): ID of the user requesting the conversation.
        proposal_id (str): ID of the proposal for which the conversation is being retrieved.

    Returns:
        (dict): A response containing the message and conversation data, along with the appropriate HTTP status code.
    """
    try:
        # Ensure that the proposal_id is a valid ObjectId
        proposal_id_obj = ObjectId(proposal_id)

        # Find the conversation in the database by user_id and proposal_id
        conversation = db.conversation.find_one({
            "user_id": ObjectId(user_id), 
            "proposal_id": proposal_id_obj
        })

        if not conversation:
            # If no conversation is found, return a 404 status code
            return make_response(
                status="error",
                message=CONVERSATION_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Convert ObjectId fields to strings for JSON serialization
        conversation["_id"] = str(conversation["_id"])
        conversation["user_id"] = str(conversation["user_id"])
        conversation["proposal_id"] = str(conversation["proposal_id"])

        # Return the conversation with a success message and 200 status code
        return make_response(
            status="success",
            message=CONVERSATION_RETRIEVED_SUCCESSFULLY,
            data=conversation,
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error in case of exceptions
        print(f"Error in retrieving conversation: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            status_code=500
        )


def save_conversation(user_id, proposal_id):
    """
    Service to save the user conversation for a specific proposal and optionally generate project vision.

    This function fetches the conversation and initiates a background save process to update the project requirements.
    If project vision does not exist, it generates the vision after saving the conversation.

    Parameters:
        user_id (str): The ID of the user.
        proposal_id (str): The ID of the proposal.

    Returns:
        (dict): A JSON response indicating the success or failure of the operation.
    """
    try:
        # Fetch the existing conversation for the given user and proposal
        conversations = db.conversation.find_one({
            "user_id": ObjectId(user_id), 
            "proposal_id": ObjectId(proposal_id)
        })

        if not conversations:
            return make_response(
                status="error",
                message=NO_CONVERSATION_FOUND,
                data=None,
                status_code=404
            )

        # Check if the proposal exists
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})
        if not proposal:
            return make_response(
                status="error",
                message="Proposal not found.",
                data=None,
                status_code=404
            )

        # Initiate background save process for the conversation
        project_description_template = db.project_description_template.find_one({})["data"]
        conversation_background_save(user_id, proposal_id, project_description_template, conversations)

        # Update the proposal status in the database
        updated_at = datetime.datetime.utcnow()
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": CLIENT_ONBOARDING_IN_PROGRESS,
                    "updated_at": updated_at
                },
            },
        )

        # Check if project vision is already available in the proposal
        if "project_vision" in proposal and proposal["project_vision"]:
            return make_response(
                status="success",
                message="Conversation saved successfully. Project vision already exists.",
                data={"project_vision": proposal["project_vision"]},
                status_code=200
            )

        # Generate the project vision only if it doesn't exist
        vision_response = generate_project_vision(user_id, proposal_id)
        if vision_response[1] != 200:
            return vision_response  # If vision generation failed, return the error

        # Return success message after saving and generating the vision
        return make_response(
            status="success",
            message=CONVERSATION_SAVED_SUCCESSFULLY + " Vision generated successfully.",
            data=vision_response[0].json.get("data"),
            status_code=200
        )

    except Exception as e:
        print(f"Error in saving conversation: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


def conversation_background_save(user_id, proposal_id, project_description_template, initial_description):
    """
    Background process to update project requirements based on the conversation.

    This function processes each section of the project description template and updates the response in the database.
    """
    updated_project_description = []
    threads = []

    # Iterate through the project description template and process each section in a new thread
    for index, description_data in enumerate(project_description_template):
        thread = threading.Thread(target=process_conversation_save_section, 
                                  args=(index, description_data, initial_description, updated_project_description, user_id, proposal_id))
        threads.append(thread)
        thread.start()

    # Wait for all threads to finish
    for thread in threads:
        thread.join()

    # Order the updated project description by index to maintain order
    ordered_project_description = [item["data"] for item in sorted(updated_project_description, key=lambda x: x["index"])]
    updated_at = datetime.datetime.utcnow()
    # Update the proposal with the updated project requirements in the database
    db.proposals.find_one_and_update(
        {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
        {
            "$set": {
                "project_requirement": ordered_project_description,
                "status": PROPOSAL_STATUS["in_progress"],
                "step": 1,
                "last_step": CLIENT_ONBOARDING_COMPLETED,
                "updated_at": updated_at
            }
        }
    )

def process_conversation_save_section(index, description_data, initial_description, updated_data, user_id, proposal_id):
    """
    Function to process a single section of the Project Requirement Template and update its response.

    Parameters:
        index (int): The index of the section in the template to maintain order.
        description_data (dict): The description data of the section (e.g., Section, Description).
        initial_description (dict): The user's initial conversation or input.
        updated_data (list): A list to hold the updated section responses.
    """
    section = description_data.get("Section")
    description = description_data.get("Description")

    if section and description:
        # Construct the question based on the section and description
        question = f"{section}: {description}"

        # Process the question and generate the response (e.g., using OpenAI or some other API)
        response = process_question(initial_description, question, user_id, proposal_id)

        # Append the processed section response to the updated data list
        updated_data.append({
            "index": index,  # Include the index to ensure the order is preserved
            "data": {
                "Section": section,
                "Description": description,
                "Response": response,
            }
        })

    return updated_data

def process_question(initial_requirements, description, user_id, proposal_id):
    """
    Processes a single question based on the initial project requirements and generates a response.
    
    Parameters:
        initial_requirements (str): The initial requirements provided by the user.
        description (str): The section description or question from the Project Requirement template.
        user_id (str): The ID of the user requesting the response.
        proposal_id (str): The ID of the proposal being processed.
    
    Returns:
        str: A generated response based on the initial requirements and the provided description.
    """
    # System context for the AI or processing logic
    system_context = (
        "You're a Product Architect and your task is to answer the Project Requirement section "
        "provided based solely on the initial project requirements received. Your response is assertive "
        "and logical, written in simple high-school level language."
    )

    # Assistant context (can be left empty or adjusted as per requirements)
    assistant_context = ""

    # Initial prompt for generating the response
    initial_prompt = (
        f"Given the initial project requirements: {initial_requirements}, "
        f"explain the following section of our Project Requirement file in detail using only the provided "
        f"initial requirements: {description}. Only use the provided initial description. If there is lacking information, "
        "say so and provide a logical answer solely based on the provided initial description."
    )

    # Call the function to generate the response (e.g., OpenAI or other API)
    response = generate_text(
        system_context, 
        assistant_context, 
        initial_prompt, 
        user_id, 
        proposal_id
    )
    
    return response


def generate_project_vision(user_id, proposal_id):
    """
    Generates the project vision for step 1 of the project.

    This function retrieves the project requirements, processes them using an external API 
    (e.g., OpenAI), and returns the generated vision based on the project data.

    Parameters:
        user_id (str): The ID of the user requesting the project vision.
        proposal_id (str): The ID of the proposal for which the vision is being generated.

    Returns:
        (dict): A JSON response with the generated project vision or an error message, along with an HTTP status code.
    """
    try:
        # Ensure the user exists before proceeding
        user = db.users.find_one({"_id": ObjectId(user_id)})

        # If the user is not found, return a 404 response
        if not user:
            return make_response(
                status="error",
                message=USER_NOT_FOUND,
                status_code=404
            )

        # Fetch the project data from the database using proposal_id and user_id
        project_data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})

        # If the project data is not found, return a 404 response
        if not project_data:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                status_code=404
            )

        # Extract the project requirements
        project_requirements = project_data.get("project_requirement", "")

        # Format the input message by replacing the placeholder with the actual project requirements
        input_message = step1InitialPrompt.replace("{project_requirements}", str(project_requirements))

        # Generate the response using an external API (e.g., OpenAI)
        data = generate_text_json(
            step1SystemContext, 
            step1AssistantContext, 
            input_message, 
            user_id, 
            proposal_id
        )

        # Parse the generated response (assuming it is a JSON string)
        data_dict = json.loads(data)

        # Return the generated data as a successful response
        return make_response(
            status="success",
            message=PROJECT_VISION_GENERATION_SUCCESS,
            data=data_dict,
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error response
        print(f"Error in generating project vision: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            status_code=500
        )
    

def get_project_vision(user_id, proposal_id):
    """
    Service to retrieve the user's project vision for step 1 of the project.

    This function ensures the existence of the user and proposal, checks if the 
    project vision exists within the proposal, and then returns the vision data.

    Parameters:
        user_id (str): The ID of the user requesting the project vision.
        proposal_id (str): The ID of the proposal for which the vision is being retrieved.

    Returns:
        (dict): A JSON response with the project vision or an error message, along with an HTTP status code.
    """
    try:
        # Ensure the user exists in the database
        user = db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            # Return a 404 error if the user is not found
            return make_response(
                status="error",
                message=USER_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Ensure the proposal exists for the given user
        proposal_ref = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})
        if not proposal_ref:
            # Return a 404 error if the proposal is not found
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Check if the proposal contains the project vision
        if "project_vision" not in proposal_ref:
            # Return a 404 error if project vision is not found within the proposal
            return make_response(
                status="error",
                message=PROJECT_VISION_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Prepare the project vision data to be returned
        data = {
            "project_vision": proposal_ref["project_vision"]
        }

        # Return the project vision along with a success message and status code 200
        return make_response(
            status="success",
            message=PROJECT_VISION_SUCCESS,
            data=data,
            status_code=200
        )

    except Exception as e:
        # Log the error and return an internal server error response with status code 500
        print(f"Error in getting project vision: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def save_project_vision(payload: SaveStep1Prompt, user_id):
    """
    Service to save the user's project vision for step 1 and generate the business vertical.

    Parameters:
        payload (dict): The payload containing project vision and proposal information.
        user_id (str): The ID of the user saving the project vision.

    Returns:
        (dict): A JSON response indicating the status of the operation, along with the project vision and business vertical data.
    """
    try:
        # Extract project vision from payload
        project_vision = payload["project_vision"]

        # Ensure the user exists
        user = db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            return make_response(
                status="error",
                message=USER_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find or create the proposal reference
        proposal_ref = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(payload["proposal_id"])})
        updated_at = datetime.datetime.utcnow()

        if proposal_ref:
            # Update existing proposal with project vision
            db.proposals.find_one_and_update(
                {"_id": proposal_ref["_id"]},
                {
                    "$set": {
                        "project_requirement": proposal_ref["project_requirement"],
                        "project_vision": project_vision,
                        "status": PROPOSAL_STATUS["in_progress"],
                        "last_step": SAVE_PROJECT_VISION_SUCCESS,
                        "step": 2,
                        "updated_at": updated_at
                    }
                }
            )
        else:
            # Insert a new proposal if none exists
            db.proposals.insert_one(
                {
                    "project_requirement": payload["project_requirement"],
                    "project_vision": project_vision,
                    "user": ObjectId(user_id),
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": SAVE_PROJECT_VISION_SUCCESS,
                    "step": 2,
                    "updated_at": updated_at
                }
            )

        # Check if the business vertical is already present
        if "business_vertical" in proposal_ref and proposal_ref["business_vertical"]:
            # If present, return the success response without re-generating
            return make_response(
                status="success",
                message=f"{SAVE_PROJECT_VISION_SUCCESS}. Business vertical already exists.",
                data={
                    "project_vision": project_vision,
                    "business_vertical": proposal_ref["business_vertical"]
                },
                status_code=200
            )

        # If business vertical is not present, generate it
        business_vertical_response = generate_business_vertical(user_id=user_id, proposal_id=payload["proposal_id"])

        # Check if the business vertical generation failed
        if business_vertical_response[1] != 200:
            return business_vertical_response  # If business vertical generation failed, return the error

        # Combine project vision and business vertical in the response
        return make_response(
            status="success",
            message=f"{SAVE_PROJECT_VISION_SUCCESS} and {BUSINESS_VERTICAL_GENERATION_SUCCESS}",
            data={
                "project_vision": project_vision,
                "business_vertical": business_vertical_response[0].json["data"]["business_vertical"]
            },
            status_code=200
        )

    except Exception as e:
        # Handle any exceptions
        print(f"Error in saving project vision or generating business vertical: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def regenerate_business_vertical_service(user_id, proposal_id):
    """
    Service to regenerate the business vertical for a specific proposal.

    Parameters:
        user_id (str): The ID of the user requesting the regeneration.
        proposal_id (str): The ID of the proposal for which to regenerate the business vertical.

    Returns:
        (dict): A JSON response indicating the success or failure of the operation.
    """
    try:
        # Ensure the user exists
        user = db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            return make_response(
                status="error",
                message=USER_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Ensure the proposal exists
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        if not proposal:
            return make_response(
                status="error",
                message="Proposal not found.",
                data=None,
                status_code=404
            )

        # Regenerate the business vertical
        business_vertical_response = generate_business_vertical(user_id=user_id, proposal_id=proposal_id)

        if business_vertical_response[1] != 200:
            return business_vertical_response  # If business vertical generation failed, return the error

        return make_response(
            status="success",
            message=BUSINESS_VERTICAL_GENERATION_SUCCESS,
            data=business_vertical_response[0].json["data"],
            status_code=200
        )

    except Exception as e:
        # Handle any exceptions
        print(f"Error in regenerating business vertical: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


def generate_business_vertical(user_id, proposal_id):
    """
    Generate the business vertical for step 2.

    Parameters:
        user_id (str): The ID of the user making the request.
        proposal_id (str): The ID of the proposal.

    Returns:
        (dict): A JSON response with the generated business vertical or an error message.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})

        # If no proposal is found, return an error
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Extract project requirements from the proposal
        project_requirement = proposal.get("project_requirement")
        if not project_requirement:
            return make_response(
                status="error",
                message="Project requirements not found.",
                data=None,
                status_code=400
            )

        # Create the input message using the predefined prompt
        input_message = step2InitialPrompt.replace("{project_requirements}", str(project_requirement))

        # Generate business vertical using the OpenAI model or a similar system
        prompt = generate_text_json(step2SystemContext, step2AssistantContext, input_message, user_id, proposal_id)

        # Parse the generated prompt result
        data = json.loads(prompt)

        # Set the current time for the updated_at field
        updated_at = datetime.datetime.utcnow()

        # Update the proposal with the generated business vertical and updated timestamp
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "business_vertical": data["data"],
                    "updated_at": updated_at  # Add updated_at timestamp
                }
            }
        )

        # Return success response with the generated business vertical
        return make_response(
            status="success",
            message=BUSINESS_VERTICAL_GENERATION_SUCCESS,
            data={"business_vertical": data["data"]},
            status_code=200
        )

    except Exception as e:
        # Log the error and return an internal server error response
        print(f"Error in generating business vertical: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            status_code=500
        )
    
def get_business_vertical(user_id, proposal_id):
    """
    Service to retrieve the business vertical for step 2.

    This function fetches the proposal based on the user ID and proposal ID,
    and then retrieves the business vertical if available.

    Parameters:
        user_id (str): The ID of the user requesting the business vertical.
        proposal_id (str): The ID of the proposal for which the business vertical is being retrieved.

    Returns:
        (dict): A JSON response with the business vertical data or an error message.
    """
    try:
        # Fetch the proposal from the database by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Retrieve the business vertical from the proposal
        business_vertical = proposal.get('business_vertical')
        if not business_vertical:
            return make_response(
                status="error",
                message=BUSINESS_VERTICAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Return success response with the business vertical data
        return make_response(
            status="success",
            message=BUSINESS_VERTICAL_FETCH_SUCCESS,
            data=json.loads(json.dumps(business_vertical, default=str)),
            status_code=200
        )

    except Exception as e:
        # Handle any unexpected errors
        print(f"Error in getting business vertical: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def update_business_vertical(payload, user_id):
    """
    Service to update the business vertical for step 2 in the proposal.

    After successfully updating the business vertical, it optionally triggers the stakeholder generation for step 3.

    Parameters:
        payload (dict): The JSON payload containing the proposal ID and the new business vertical data.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the status of the update and stakeholder generation.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(payload["proposal_id"])})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Set the current time for the updated_at field
        updated_at = datetime.datetime.utcnow()

        # Update the proposal with the new business vertical and status information
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(payload["proposal_id"])},
            {
                "$set": {
                    "business_vertical": payload["business_vertical"],
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": BUSINESS_VERTICAL_UPDATE_SUCCESS,
                    "step": 3,
                    "updated_at": updated_at  # Set updated_at field
                }
            }
        )

        # Check if the stakeholders are already present
        if "stake_holders" in proposal and proposal["stake_holders"]:
            # If stakeholders are already present, skip generation and return success message
            return make_response(
                status="success",
                message=f"{BUSINESS_VERTICAL_UPDATE_SUCCESS}. Stakeholders already exist.",
                data={
                    "business_vertical": payload["business_vertical"],
                    "stake_holders": proposal["stake_holders"]
                },
                status_code=200
            )

        # Call the stakeholder generation process after successfully updating the business vertical
        stakeholder_response = generate_stakeholders(proposal_id=payload["proposal_id"], user_id=user_id)

        # If the stakeholder generation fails, return its error response
        if stakeholder_response[1] != 200:
            return stakeholder_response  # If the stakeholder generation failed, return its error response

        # Return success response with the business vertical and generated stakeholders
        return make_response(
            status="success",
            message=f"{BUSINESS_VERTICAL_UPDATE_SUCCESS} and {STAKEHOLDER_GENERATION_SUCCESS}",
            data={
                "business_vertical": payload["business_vertical"],
                "stake_holders": stakeholder_response[0].json["data"]["stake_holders"]
            },
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error in case of exceptions
        print(f"Error in updating business vertical: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )
    
def regenerate_stakeholders_service(user_id, proposal_id):
    """
    Service to regenerate the stakeholders for a specific proposal.

    Parameters:
        user_id (str): The ID of the user requesting the regeneration.
        proposal_id (str): The ID of the proposal for which to regenerate the stakeholders.

    Returns:
        (dict): A JSON response indicating the success or failure of the operation.
    """
    try:
        # Ensure the user exists
        user = db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            return make_response(
                status="error",
                message=USER_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Ensure the proposal exists
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Regenerate the stakeholders
        stakeholder_response = generate_stakeholders(proposal_id=proposal_id, user_id=user_id)

        if stakeholder_response[1] != 200:
            return stakeholder_response  # If the stakeholder generation failed, return the error

        return make_response(
            status="success",
            message=STAKEHOLDER_GENERATION_SUCCESS,
            data=stakeholder_response[0].json["data"],
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error in case of exceptions
        print(f"Error in regenerating stakeholders: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


def generate_stakeholders(proposal_id, user_id):
    """
    Service to generate stakeholders for step 3 in the proposal process.

    This function retrieves the project requirements and conversation history,
    generates the stakeholders using an external system (e.g., OpenAI), and updates the proposal.

    Parameters:
        proposal_id (str): The ID of the proposal for which stakeholders are being generated.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the success or failure of the generation process.
    """
    try:
        # Retrieve the project requirements from the proposal
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})

        # If the proposal is not found, return an error response
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        project_requirement = proposal.get("project_requirement")
        if not project_requirement:
            return make_response(
                status="error",
                message="Project requirements not found.",
                data=None,
                status_code=400
            )

        # Retrieve the conversation related to the proposal
        conversation = db.conversation.find_one({"user_id": ObjectId(user_id), "proposal_id": ObjectId(proposal_id)})

        # If the conversation is not found, return an error response
        if not conversation:
            return make_response(
                status="error",
                message=CONVERSATION_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Prepare the input message by replacing placeholders with actual data
        input_message = step3InitialPrompt.replace("{project_requirements}", str(project_requirement))
        input_message = input_message.replace("{conversation_thread}", str(conversation["data"]))

        # Generate stakeholders using the external system (e.g., OpenAI)
        prompt = generate_text_json(step3SystemContext, step3AssistantContext, input_message, user_id, proposal_id)
        stakeholders_data = json.loads(prompt)

        # Set the current time for the updated_at field
        updated_at = datetime.datetime.utcnow()

        # Update the proposal with the generated stakeholders and timestamps
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "stake_holders": stakeholders_data["data"],
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": STAKEHOLDER_GENERATION_SUCCESS,
                    "updated_at": updated_at  # Set updated_at field
                }
            }
        )

        # Return success response with the generated stakeholders
        return make_response(
            status="success",
            message=STAKEHOLDER_GENERATION_SUCCESS,
            data={"stake_holders": stakeholders_data["data"]},
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error response
        print(f"Error in generating stakeholders: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def get_stakeholders(proposal_id, user_id):
    """
    Service to retrieve stakeholders for step 3 in the proposal process.

    Parameters:
        proposal_id (str): The ID of the proposal for which stakeholders are being retrieved.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the success or failure of the retrieval process.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Retrieve the stakeholders from the proposal
        stakeholders = proposal.get("stake_holders")
        if not stakeholders:
            return make_response(
                status="error",
                message=STAKEHOLDERS_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Return success response with the stakeholders data
        return make_response(
            status="success",
            message=STAKEHOLDERS_FETCH_SUCCESS,
            data={"stake_holders": stakeholders},
            status_code=200
        )

    except Exception as e:
        # Log the error and return an internal server error response
        print(f"Error in getting stakeholders: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def update_stakeholders(payload, user_id):
    """
    Service to update stakeholders for step 3 in the proposal process and then generate revenue model for step 4.

    Parameters:
        payload (dict): The JSON payload containing the proposal ID and the new stakeholders data.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the status of the update and revenue model generation operation.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(payload["proposal_id"])})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Set the current time for the updated_at field
        updated_at = datetime.datetime.utcnow()

        # Update the proposal with the new stakeholders and the current step
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(payload["proposal_id"])},
            {
                "$set": {
                    "stake_holders": payload["stake_holders"],
                    "last_step": STAKEHOLDER_GENERATION_SUCCESS,
                    "step": 3,
                    "updated_at": updated_at  # Set updated_at field
                }
            }
        )

        # After updating stakeholders, generate the revenue model for step 4
        revenue_response = generate_revenue_model(proposal_id=payload["proposal_id"], user_id=user_id)

        # If the revenue model generation fails, return its error response
        if revenue_response[1] != 200:
            return revenue_response

        # Return success response after updating stakeholders and generating the revenue model
        return make_response(
            status="success",
            message=f"{STAKEHOLDERS_UPDATED_SUCCESS} and {REVENUE_MODEL_GENERATION_SUCCESS}",
            data={
                "stake_holders": payload["stake_holders"],
                "revenue_model": revenue_response[0].json["data"]["revenue_model"]
            },
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error in case of exceptions
        print(f"Error in updating stakeholders and generating revenue model: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


def generate_revenue_model(proposal_id, user_id):
    """
    Service to generate the revenue model for step 4 of the proposal process.

    This function retrieves the project requirements and stakeholders, generates
    the revenue model using an external system (e.g., OpenAI), and updates the proposal.

    Parameters:
        proposal_id (str): The ID of the proposal for which the revenue model is being generated.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the success or failure of the revenue model generation.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})
        print("🚀 ~ proposal:", proposal["stake_holders"])

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Prepare the input message for generating the revenue model by replacing placeholders
        input_message = step4InitialPrompt.replace("{project_requirements}", str(proposal["project_requirement"]))
        input_message = input_message.replace("[{primary_stakeholders}]", str(proposal["stake_holders"]))

        # Generate the revenue model using the external system (e.g., OpenAI)
        prompt = generate_text_json(step4SystemContext, step4AssistantContext, input_message, user_id, proposal_id)

        # Parse the generated response
        data = json.loads(prompt)

        # Set the current time for the updated_at field
        updated_at = datetime.datetime.utcnow()

        # Update the proposal with the generated revenue model and status information
        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "revenue_model": data["data"],
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": REVENUE_MODEL_GENERATION_SUCCESS,
                    "step": 4,
                    "updated_at": updated_at  # Set updated_at field
                }
            }
        )

        # Return success response with the generated revenue model
        return make_response(
            status="success",
            message=REVENUE_MODEL_GENERATION_SUCCESS,
            data={"revenue_model": data["data"]},
            status_code=200
        )

    except Exception as e:
        # Log and return an internal server error response in case of exceptions
        print(f"Error in generating revenue model: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def get_revenue_model(proposal_id, user_id):
    """
    Service to retrieve the revenue model for step 4 in the proposal process.

    Parameters:
        proposal_id (str): The ID of the proposal for which the revenue model is being retrieved.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response indicating the success or failure of the retrieval process.
    """
    try:
        # Fetch the proposal by user ID and proposal ID
        proposal = db.proposals.find_one({"user": ObjectId(user_id), "_id": ObjectId(proposal_id)})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Retrieve the revenue model from the proposal
        revenue_model = proposal.get("revenue_model")
        if not revenue_model:
            return make_response(
                status="error",
                message=REVENUE_MODEL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Return success response with the revenue model data
        return make_response(
            status="success",
            message=REVENUE_MODEL_FETCH_SUCCESS,
            data={"revenue_model": revenue_model},
            status_code=200
        )

    except Exception as e:
        # Log the error and return an internal server error response
        print(f"Error in getting revenue model: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )
    
def fetch_epics_by_proposal(proposal_id, user_id):
    """
    Service to fetch all epics for a given proposal based on the proposal ID and user ID.

    Parameters:
        proposal_id (str): The ID of the proposal.
        user_id (ObjectId): The ID of the user.

    Returns:
        (dict): A JSON response with epics, stakeholder, status, step, and a flag epicsIsPresent.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not data:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                status_code=404
            )

        # Get epics from the data and check if it's present
        epics = data.get("epics", [])

        # Prepare the response data
        response_data = {
            "epicsIsPresent": len(epics) > 0,  # True if epics list has elements
            "status": data.get("status", ""),
            "step": data.get("step", ""),
        }

        return make_response(
            status="success",
            message=EPICS_FETCHED_SUCCESS,
            data=response_data,
            status_code=200
        )

    except Exception as e:
        print(f"Error fetching epics: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            status_code=500
        )

def generate_epics(proposal_id, user_id, user_email):
    """
    Service to generate epics for step 5 of the proposal process.

    Parameters:
        proposal_id (str): The ID of the proposal for which epics are being generated.
        user_id (str): The ID of the user making the request.
        user_email (str): The email of the user making the request.

    Returns:
        (dict): A JSON response indicating the success or failure of the epic generation operation.
    """
    try:
        # Fetch the proposal and conversation by user ID and proposal ID
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        conversation = db.conversation.find_one({"user_id": ObjectId(user_id), "proposal_id": ObjectId(proposal_id)})

        # If the proposal is not found, return an appropriate error message
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Update the proposal status to "in progress" for generating epics
        db.proposals.find_one_and_update(
            {"_id": ObjectId(proposal_id)},
            {"$set": {"status": PROPOSAL_STATUS["in_progress"], "last_step": EPICS_GENERATION_STARTED}},
        )

        # Start the process to generate epics synchronously
        result = run_async(partial(background_save_epics, proposal, conversation, proposal_id, user_email, user_id))

        # Return success response after epics are generated
        if result:
            return make_response(
                status="success",
                message=EPICS_GENERATION_SUCCESS,
                data=None,
                status_code=200
            )
        else:
            return make_response(
                status="error",
                message=EPICS_GENERATION_FAILED,
                data=None,
                status_code=500
            )

    except Exception as e:
        # Log and return an internal server error response in case of exceptions
        print(f"Error in generating epics: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


async def background_save_epics(proposal, conversation, proposal_id, user_email, user_id):
    """
    Background process to generate epics and stories for the proposal.

    Parameters:
        proposal (dict): The proposal data.
        conversation (dict): The conversation data.
        proposal_id (str): The ID of the proposal.
        user_email (str): The email of the user.
        user_id (str): The ID of the user.

    Returns:
        (list): The generated epics, each with a unique epic_id.
    """
    try:
        system_context = define_epics_system_context
        assistant_context = define_epics_assistant_context

        input_message = define_epics_initial_prompt.replace("{initial_requirements}", str(conversation["data"]))
        input_message = input_message.replace("{project_requirements}", str(proposal["project_vision"]))

        result = []

        # Generate epics for each stakeholder
        for stakeholder in proposal["stake_holders"]:
            temp_input_message = input_message.replace("{stakeholder}", stakeholder)

            # Generate the epics using the AI model
            prompt = generate_text_json_with_4o(system_context, assistant_context, temp_input_message, user_id, proposal_id)
            epics_data = json.loads(prompt)["Epics"]

            # Add a unique epic_id (MongoDB ObjectId) to each epic
            for epic in epics_data:
                epic["id"] = str(ObjectId())  # Generate a unique ObjectId for each epic

            result.append({"stakeholder": stakeholder, "data": epics_data})

        # Update the proposal with the generated epics, including their unique epic_id
        updated_at = datetime.datetime.utcnow()
        db.proposals.find_one_and_update(
            {"_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "epics": result,
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": EPICS_GENERATED,
                    "updated_at": updated_at
                }
            },
        )

        return result

    except Exception as e:
        print(f"Error in saving epics: {e}")
        return None

# def run_async(async_fn):
#     """
#     Function to run an asynchronous function in a synchronous context by manually
#     creating and running an asyncio event loop.

#     Parameters:
#         async_fn (coroutine function): The asynchronous function to run.

#     Returns:
#         Any: The result returned by the asynchronous function.
#     """
#     try:
#         # Create a new event loop for asyncio
#         loop = asyncio.new_event_loop()

#         # Set the new event loop as the current event loop
#         asyncio.set_event_loop(loop)

#         # Run the asynchronous function and block until it completes
#         result = loop.run_until_complete(async_fn())

#         # Close the event loop after the async function completes
#         loop.close()

#         # Return the result of the async function
#         return result

#     except Exception as e:
#         print(f"Error in running async function: {e}")
#         return None
    
def run_async(async_fn):
    # Start the asyncio event loop and run the async function
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(async_fn())
    loop.close()
    return result

def add_new_epic(payload, user_id):
    """
    Service to add a new epic for a specific stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, title, and description.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message and the new epic, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})
        
        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )
        
        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                epics_list = stakeholder_data["data"]
                
                # Generate the new epic_id based on the length of current epics
                new_epic_id = f"E{len(epics_list) + 1:03d}"

                # Create the new epic with the given title, description, and generated epic_id
                new_epic = {
                    "epic_id": new_epic_id,
                    "title": payload["title"],
                    "description": payload["description"],
                    "id":str(ObjectId())
                }
                
                # Append the new epic to the stakeholder's epic list
                epics_list.append(new_epic)

                # Update the proposal in the database with the new epic
                db.proposals.find_one_and_update(
                    {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                    {
                        "$set": {
                            "epics.$[stakeholder].data": epics_list,
                            "updated_at": datetime.datetime.utcnow()
                        }
                    },
                    array_filters=[{"stakeholder.stakeholder": payload["stakeholder"]}]
                )

                # Return a success response with the new epic details
                return make_response(
                    status="success",
                    message=EPIC_ADDED_SUCCESS,
                    data={"epic": new_epic},
                    status_code=200
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in adding epic: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def update_existing_epic(payload, user_id):
    """
    Service to update an existing epic for a specific stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, title, and description.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message and the updated epic, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})
        
        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )
        
        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                epics_list = stakeholder_data["data"]

                # Find the epic to update by id
                for epic in epics_list:
                    if epic["id"] == payload["id"]:
                        # Update the title and description of the selected epic
                        epic["title"] = payload["title"]
                        epic["description"] = payload["description"]
                        
                        # Update the proposal in the database with the updated epic
                        db.proposals.find_one_and_update(
                            {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                            {
                                "$set": {
                                    "epics.$[stakeholder].data": epics_list,
                                    "updated_at": datetime.datetime.utcnow()
                                }
                            },
                            array_filters=[{"stakeholder.stakeholder": payload["stakeholder"]}]
                        )

                        # Return a success response with the updated epic details
                        return make_response(
                            status="success",
                            message=EPIC_UPDATED_SUCCESS,
                            data={"epic": epic},
                            status_code=200
                        )

                # If id is not found, return an error message
                return make_response(
                    status="error",
                    message=EPIC_NOT_FOUND,
                    data=None,
                    status_code=404
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in updating epic: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def delete_epic_by_id(payload, user_id):
    """
    Service to delete an epic for a specific stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, and epic_id.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message, or an error message if not found.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                epics_list = stakeholder_data["data"]

                # Find the index of the epic to delete by id
                epic_to_delete = None
                for epic in epics_list:
                    if epic["id"] == payload["id"]:
                        epic_to_delete = epic
                        break

                # If the epic is found, remove it from the list
                if epic_to_delete:
                    epics_list.remove(epic_to_delete)

                    # Update the proposal in the database with the updated epic list
                    db.proposals.find_one_and_update(
                        {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                        {
                            "$set": {
                                "epics.$[stakeholder].data": epics_list,
                                "updated_at": datetime.datetime.utcnow()
                            }
                        },
                        array_filters=[{"stakeholder.stakeholder": payload["stakeholder"]}]
                    )

                    # Return a success response indicating the epic was deleted
                    return make_response(
                        status="success",
                        message=EPIC_DELETED_SUCCESS,
                        data=None,
                        status_code=200
                    )

                # If the id is not found, return an error message
                return make_response(
                    status="error",
                    message=EPIC_NOT_FOUND,
                    data=None,
                    status_code=404
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in deleting epic: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )
    
def generate_story_basedon_epics(proposal_id, user_id, user_email):
    """
    Generate stories based on epics for a specific proposal.

    Parameters:
        proposal_id (str): The ID of the proposal for which stories are being generated.
        user_id (str): The ID of the user making the request.
        user_email (str): The email of the user making the request.

    Returns:
        dict: A JSON response indicating the success or failure of the story generation operation.
    """
    try:
        # Fetch proposal and conversation
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": user_id})
        conversation = db.conversation.find_one({"user_id": user_id, "proposal_id": ObjectId(proposal_id)})

        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Update proposal status to 'in-progress'
        db.proposals.find_one_and_update(
            {"_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": GENERATING_STORIES,
                    "updated_at": datetime.datetime.utcnow()
                }
            },
        )

        # Fetch user data
        user_data = db.users.find_one({"_id": ObjectId(user_id)})

        # Run the async function to generate stories using ThreadPoolExecutor
        async_fn = partial(
            background_save_stories,
            proposal,
            conversation,
            proposal_id,
            user_email,
            user_data,
            user_id
        )

        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(run_async, async_fn)
            results = future.result()  # Get the result of the background task

        # Return success response
        return make_response(
            status="success",
            message=None,
            data=results,
            status_code=200
        )

    except Exception as e:
        print(f"Error in generating stories: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

async def background_save_stories(data, conversation, proposal_id, user_email, user_data, user_id):
    """
    Background task to generate stories for each epic.

    Parameters:
        data (dict): The proposal data.
        conversation (dict): The conversation data.
        proposal_id (str): The ID of the proposal.
        user_email (str): The email of the user.
        user_data (dict): The user data.
        user_id (str): The ID of the user.

    Returns:
        list: A list of generated stories for each stakeholder's epics.
    """
    try:
        tasks = [
            process_satckholder_stories(data, stakeholder, user_data, proposal_id, user_id)
            for stakeholder in data["stake_holders"]
        ]
        results = await asyncio.gather(*tasks)
        return results

    except Exception as e:
        print(f"Error in generating stories for epics: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )


async def process_satckholder_stories(data, stakeholder_name, user_data, proposal_id, user_id):
    """
    Define user stories for each epic within the stakeholder.

    Parameters:
        data (dict): The proposal data.
        stakeholder_name (str): The name of the stakeholder.
        user_data (dict): The user data.
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user.

    Returns:
        dict: A stakeholder with its updated epics including user stories.
    """
    try:
        # Define contexts for story and task generation
        system_context = user_stories_system_context
        assistant_context = user_stories_assistant_context

        input_message = user_stories_initial_prompt.replace(
            "{client_requirements}", str(data["project_requirement"])
        ).replace(
            "{project_requirements}", str(data["project_vision"])
        ).replace(
            "{list_of_stakeholders}", str(data["stake_holders"])
        )

        # Find the stakeholder
        stakeholder = next(
            (item for item in data["epics"] if item.get("stakeholder") == stakeholder_name),
            None
        )

        if not stakeholder:
            print(f"Stakeholder '{stakeholder_name}' not found in epics")
            return make_response(
                status="error",
                message=STAKEHOLDER_NOT_FOUND,
                data=None,
                status_code=404
            )

        input_message = input_message.replace("{stakeholder}", stakeholder["stakeholder"])

        # Generate stories for each epic within the stakeholder
        tasks = [
            process_epic(
                system_context, assistant_context, input_message, epics_data, stakeholder, data, user_data, proposal_id, user_id
            ) for epics_data in stakeholder["data"]
        ]

        # Await the result of the async tasks and update the proposal with the generated stories
        stakeholder["data"] = await asyncio.gather(*tasks)

        db.proposals.update_one(
            {"_id": data["_id"], "epics.stakeholder": stakeholder_name},
            {
                "$set": {
                    "epics.$.data": stakeholder["data"],
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": GENERATING_STORIES,
                    "updated_at": datetime.datetime.utcnow()
                }
            }
        )

        return stakeholder

    except Exception as e:
        print(f"Error in defining user stories for stakeholder '{stakeholder_name}': {e}")
        raise e


async def process_epic(system_context, assistant_context, input_message, epics_data, stakeholder, data, user_data, proposal_id, user_id):
    """
    Process and generate user stories for a specific epic.

    Parameters:
        system_context (str): The system context for story generation.
        assistant_context (str): The assistant context for story generation.
        input_message (str): The input message template.
        epics_data (dict): The epic data.
        stakeholder (dict): The stakeholder data.
        data (dict): The proposal data.
        user_data (dict): The user data.
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user.

    Returns:
        dict: The epic with generated user stories.
    """
    try:
        # Prepare the input message for the specific epic
        temp_input_message = input_message.replace("{epic}", epics_data["description"])
        prompt_response = await async_generate_text_json_with_4o(
            system_context, assistant_context, temp_input_message, user_id, proposal_id
        )
        prompt = json.loads(prompt_response)

        # Iterate through each story and assign a unique MongoDB ObjectId
        stories = prompt["UserStories"]
        for story in stories:
            story["id"] = str(ObjectId())  # Add a unique MongoDB ObjectId for each story

        # Assign the generated user stories (with unique story_id) to the epic
        epics_data["user_stories"] = stories

        return epics_data

    except Exception as e:
        print(f"Error in processing epic '{epics_data['description']}': {e}")
        raise e


async def async_generate_text_json_with_4o(system_context, assistant_context, input_message,user_id, praposal_id):
    loop = asyncio.get_event_loop()
 
    with concurrent.futures.ThreadPoolExecutor() as executor:
        result = await loop.run_in_executor(executor, generate_text_json_with_4o, system_context, assistant_context, input_message,user_id, praposal_id)
    return result

def generate_tasks_basedon_stories(proposal_id, user_id, user_email):
    """
    Service to generate tasks based on stories for a specific proposal.

    Parameters:
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user making the request.
        user_email (str): The email of the user.

    Returns:
        dict: A JSON response indicating the success or failure of the task generation.
    """
    try:
        # Fetch proposal and conversation data for the given proposal and user
        data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        conversation = db.conversation.find_one(
            {"user_id": ObjectId(user_id), "proposal_id": ObjectId(proposal_id)}
        )

        # If proposal not found, return an error response
        if not data:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Update the proposal status to "in progress"
        db.proposals.find_one_and_update(
            {"_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": GENERATING_TASKS,
                }
            },
        )

        # Fetch user data for task generation
        user_data = db.users.find_one({"_id": ObjectId(user_id)})

        # Start background task generation process using threading
        async_fn = partial(
            background_process_for_task,
            data,
            conversation,
            proposal_id,
            user_email,
            user_data,
            user_id
        )

        # Use ThreadPoolExecutor to run the async function in a new thread
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(run_async, async_fn)
            results = future.result()

        # Return success response after task generation
        return make_response(
            status="success",
            message=TASKS_GENERATED_SUCCESS,
            data=results,
            status_code=200
        )

    except Exception as e:
        # Handle any exceptions and return an internal server error response
        print(f"Error in generating tasks: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

async def background_process_for_task(data, conversation, proposal_id, user_email, user_data, user_id):
    """
    Background process to generate tasks for epics in the proposal.

    Parameters:
        data (dict): The proposal data.
        conversation (dict): The conversation data.
        proposal_id (str): The ID of the proposal.
        user_email (str): The email of the user.
        user_data (dict): The user data.
        user_id (str): The ID of the user.

    Returns:
        list: The generated tasks for each epic.
    """
    try:
        # Process task generation for each stakeholder
        tasks = [
            process_stakeholder_for_task(data, stakeholder, user_data, proposal_id, user_id)
            for stakeholder in data["stake_holders"]
        ]
        # Await all stakeholder task processing asynchronously
        results = await asyncio.gather(*tasks)
        return results
    except Exception as e:
        # Handle exceptions in background task processing
        print(f"Error in background task processing: {e}")
        raise e

async def process_stakeholder_for_task(data, stakeholder_name, user_data, proposal_id, user_id):
    """
    Process task generation for a stakeholder's epics.

    Parameters:
        data (dict): The proposal data.
        stakeholder_name (str): The name of the stakeholder.
        user_data (dict): The user data.
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user.

    Returns:
        dict: The processed tasks for the stakeholder's epics.
    """
    try:
        # Build input message for the assistant context
        input_message = user_stories_initial_prompt.replace(
            "{client_requirements}", str(data["project_requirement"])
        ).replace(
            "{project_requirements}", str(data["project_vision"])
        ).replace(
            "{list_of_stakeholders}", str(data["stake_holders"])
        )

        system_context_task = task_definition_system_context
        assistant_context_task = task_definition_assistant_context

        # Find the stakeholder by name
        stakeholder = next(
            (item for item in data["epics"] if item.get("stakeholder") == stakeholder_name),
            None
        )

        # If stakeholder not found, skip processing
        if not stakeholder:
            print(f"Stakeholder '{stakeholder_name}' not found in epics")
            return

        # Process tasks for each epic under the stakeholder
        input_message_task = task_definition_initial_prompt.replace("{stakeholder}", stakeholder["stakeholder"])
        tasks = [
            process_epic_for_task(
                system_context_task, assistant_context_task, input_message_task, epics_data,
                stakeholder, data, user_data, proposal_id, user_id
            )
            for epics_data in stakeholder["data"]
        ]

        # Gather and update the tasks asynchronously
        stakeholder["data"] = await asyncio.gather(*tasks)

        # Update the proposal with the processed tasks
        db.proposals.update_one(
            {"_id": ObjectId(proposal_id), "epics.stakeholder": stakeholder_name},
            {
                "$set": {
                    "epics.$.data": stakeholder["data"],
                    "status": PROPOSAL_STATUS["in_progress"],
                    "last_step": GENERATING_TASKS,
                }
            }
        )
        return stakeholder

    except Exception as e:
        # Handle errors during task processing for stakeholder
        print(f"Error in processing tasks for stakeholder '{stakeholder_name}': {e}")
        raise e

async def process_epic_for_task(system_context_task, assistant_context_task, input_message_task, epics_data, stakeholder, data, user_data, proposal_id, user_id):
    """
    Process task generation for an epic.

    Parameters:
        system_context_task (str): System context for task generation.
        assistant_context_task (str): Assistant context for task generation.
        input_message_task (str): The input message template.
        epics_data (dict): The epic data.
        stakeholder (dict): The stakeholder data.
        data (dict): The proposal data.
        user_data (dict): The user data.
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user.

    Returns:
        dict: The epic with generated tasks.
    """
    try:
        # Process each story within the epic and generate tasks
        tasks = [
            process_story_for_task(epics_data, system_context_task, assistant_context_task, input_message_task, story, stakeholder, data, user_data, proposal_id, user_id)
            for story in epics_data["user_stories"]
        ]

        # Gather and update the stories asynchronously
        updated_user_stories = await asyncio.gather(*tasks)
        epics_data["user_stories"] = updated_user_stories

        return epics_data

    except Exception as e:
        # Handle errors during task processing for an epic
        print(f"Error in processing tasks for epic '{epics_data['description']}': {e}")
        raise e

async def process_story_for_task(epics_data, system_context_task, assistant_context_task, input_message_task, story, stakeholder, data, user_data, proposal_id, user_id):
    """
    Process task generation for a story within an epic.

    Parameters:
        epics_data (dict): The epic data.
        system_context_task (str): System context for task generation.
        assistant_context_task (str): Assistant context for task generation.
        input_message_task (str): The input message template.
        story (dict): The story data.
        stakeholder (dict): The stakeholder data.
        data (dict): The proposal data.
        user_data (dict): The user data.
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user.

    Returns:
        dict: The story with generated tasks.
    """
    try:
        # Replace placeholders in the input message with story details
        story_input_message = input_message_task.replace("{story}", str(story["description"])).replace(
            "{acceptance_criteria}", str(story["acceptance_criteria"])
        )

        # Generate task using async function
        prompt_response = await async_generate_text_json(
            system_context_task, assistant_context_task, story_input_message, user_id, proposal_id
        )
        prompt = json.loads(prompt_response)

        # Analyze the complexity of the tasks and estimate hours and cost
        story["tasks"] = await async_complexity_analysis(
            prompt["tasks"], story["title"], epics_data["title"], stakeholder["stakeholder"],
            user_data, user_id, proposal_id
        )

        return story

    except Exception as e:
        # Handle errors during task processing for a story
        print(f"Error in processing tasks for story '{story['description']}': {e}")
        raise e

async def async_calculate_complexity_hours(userData, complexity_level):
    """
    Calculate the complexity hours and cost based on the complexity level and user settings.

    Parameters:
        userData (dict): The user data containing complexity settings and hourly rates.
        complexity_level (str): The determined complexity level for the task.

    Returns:
        dict: Estimated hours and cost for the task.
    """
    try:
        # Validate the required settings are available in userData
        if not userData or "settings" not in userData or "level_of_complexities" not in userData["settings"] or "hourly_rate" not in userData["settings"]:
            raise ValueError("Invalid userData structure. Missing settings, level_of_complexities, or hourly_rate.")

        # Retrieve complexity levels and hourly rates from user settings
        level_of_complexities = userData["settings"]["level_of_complexities"]
        hourly_rate = userData["settings"]["hourly_rate"]

        # Determine the complexity level
        complexity_level_str = determine_complexity_level(complexity_level)

        # Validate the complexity level exists in user settings
        if complexity_level_str not in level_of_complexities:
            raise ValueError(f"Complexity level '{complexity_level_str}' not found in level_of_complexities.")

        # Calculate estimated hours and cost
        estimated_hours = level_of_complexities[complexity_level_str]
        cost = estimated_hours * hourly_rate
        return {"cost": cost, "estimated_hours": estimated_hours}

    except ValueError as ve:
        print(f"Validation error: {ve}")
        raise ve
    except Exception as e:
        print("Error in calculating complexity hours", e)
        raise e

async def async_complexity_analysis(tasks, story, epic, stakeholder, userData, user_id, proposal_id):
    """
    Analyze the complexity of tasks and estimate the hours and cost for each task.

    Parameters:
        tasks (list): List of tasks to analyze.
        story (str): The story title.
        epic (str): The epic title.
        stakeholder (str): The stakeholder name.
        userData (dict): The user data containing complexity settings.
        user_id (str): The ID of the user.
        proposal_id (str): The ID of the proposal.

    Returns:
        list: Tasks with estimated hours, cost, and complexity analysis.
    """
    try:
        if not tasks or not isinstance(tasks, list):
            raise ValueError("Tasks must be a non-empty list.")

        system_context = task_complexity_system_context.replace(
            "{task_complexity_options}", task_complexity_options
        )
        assistant_context = task_complexity_assistant_context
        initial_prompt = task_complexity_initial_prompt.replace("{tasks}", str(tasks))
        initial_prompt = initial_prompt.replace("{story}", story)
        initial_prompt = initial_prompt.replace("{epic}", epic)
        initial_prompt = initial_prompt.replace("{stakeholder}", stakeholder)

        result = []

        # Process each task and calculate complexity analysis
        async def async_process_task(task):
            temp_initial_prompt = initial_prompt
            prompt = await async_generate_prompt_response(system_context, assistant_context, temp_initial_prompt, user_id, proposal_id)
            estimate = await async_calculate_complexity_hours(userData, prompt)
            task["estimated_hours"] = estimate["estimated_hours"]
            task["cost"] = estimate["cost"]
            task["complexity"] = prompt
            task["id"] = str(ObjectId())  # Add unique task ID
            result.append(task)

        # Run all tasks concurrently
        await asyncio.gather(*[async_process_task(task) for task in tasks])

        return result

    except Exception as e:
        print(f"Error in complexity analysis: {e}")
        raise e

# Helper function to determine complexity level
def determine_complexity_level(complexity_level):
    """
    Helper function to determine the complexity level string based on the complexity level.

    Parameters:
        complexity_level (str): The raw complexity level value.

    Returns:
        str: The standardized complexity level string.
    """
    if "Very_Simple" in complexity_level:
        return "Very_Simple"
    elif "Simple" in complexity_level:
        return "Simple"
    elif "Medium" in complexity_level:
        return "Medium"
    elif "Complex" in complexity_level and "Very_Complex" not in complexity_level:
        return "Complex"
    elif "Very_Complex" in complexity_level:
        return "Very_Complex"
    else:
        return "Medium"  # Default complexity level


async def async_generate_text_json(system_context, assistant_context, input_message,user_id, praposal_id):
    loop = asyncio.get_event_loop()
 
    with concurrent.futures.ThreadPoolExecutor() as executor:
        result = await loop.run_in_executor(executor, generate_text_json, system_context, assistant_context, input_message,user_id, praposal_id)
    return result

async def async_generate_prompt_response(system_context, assistant_context, input_message, user_id, praposal_id):
    loop = asyncio.get_event_loop()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        result = await loop.run_in_executor(executor, generate_text, system_context, assistant_context, input_message,user_id, praposal_id)
    return result

def add_new_story(payload, user_id):
    """
    Service to add a new story for a specific epic within a stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, title, description, and acceptance_criteria.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message and the new story, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == payload["epic_id"]:
                        print("🚀 ~ epic:", epic)
                        # Get the current list of stories
                        stories_list = epic.get("user_stories", [])
                        
                        # Generate the new story_id based on the length of current stories
                        new_story_id = f"S{len(stories_list) + 1:03d}"

                        # Create the new story with the given title, description, and acceptance_criteria
                        new_story = {
                            "story_id": new_story_id,
                            "title": payload["title"],
                            "description": payload["description"],
                            "acceptance_criteria": payload["acceptance_criteria"],  # List format
                            "id": str(ObjectId())
                        }

                        # Append the new story to the epic's stories list
                        stories_list.append(new_story)
                        print("🚀 ~ stories_list:", stories_list)

                        # Update the epic in the proposal with the new story
                        db.proposals.find_one_and_update(
                            {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                            {
                                "$set": {
                                    "epics.$[stakeholder].data.$[epic].user_stories": stories_list,
                                    "updated_at": datetime.datetime.utcnow()
                                }
                            },
                            array_filters=[
                                {"stakeholder.stakeholder": payload["stakeholder"]},
                                {"epic.id": payload["epic_id"]}
                            ]
                        )

                        # Return a success response with the new story details
                        return make_response(
                            status="success",
                            message=STORY_ADDED_SUCCESS,
                            data={"story": new_story},
                            status_code=200
                        )

        # If the stakeholder or epic is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_OR_EPIC_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in adding story: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def add_new_task(payload, user_id):
    """
    Service to add a new task for a specific story within an epic for a stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, story_id, description, and complexity.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message and the new task, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == payload["epic_id"]:
                        # Find the specific story in the epic
                        for story in epic["user_stories"]:
                            if story["id"] == payload["story_id"]:
                                # Get the current list of tasks
                                tasks_list = story.get("tasks", [])

                                # Generate the new task_id based on the length of current tasks
                                new_task_id = f"T{len(tasks_list) + 1:03d}"

                                # Fetch user data for complexity analysis
                                user_data = db.users.find_one({"_id": ObjectId(user_id)})
                                print('asf')
                                # Calculate the estimated hours and cost using the existing complexity function
                                complexity_result = async_calculate_complexity_hours_for_add_task(user_data, payload["complexity"])

                                # Create the new task with the given description and complexity
                                new_task = {
                                    "task_id": new_task_id,
                                    "description": payload["description"],
                                    "complexity": payload["complexity"],
                                    "estimated_hours": complexity_result["estimated_hours"],
                                    "cost": complexity_result["cost"],
                                    "id": str(ObjectId())  # Unique MongoDB Object ID
                                }

                                # Append the new task to the story's tasks list
                                tasks_list.append(new_task)

                                # Update the story in the proposal with the new task
                                db.proposals.find_one_and_update(
                                    {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                                    {
                                        "$set": {
                                            "epics.$[stakeholder].data.$[epic].user_stories.$[story].tasks": tasks_list,
                                            "updated_at": datetime.datetime.utcnow()
                                        }
                                    },
                                    array_filters=[
                                        {"stakeholder.stakeholder": payload["stakeholder"]},
                                        {"epic.id": payload["epic_id"]},
                                        {"story.id": payload["story_id"]}
                                    ]
                                )

                                # Return a success response with the new task details
                                return make_response(
                                    status="success",
                                    message=TASK_ADDED_SUCCESS,
                                    data={"task": new_task},
                                    status_code=200
                                )

        # If the stakeholder, epic, or story is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_EPIC_OR_STORY_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in adding task: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def async_calculate_complexity_hours_for_add_task(userData, complexity_level):
    """
    Calculate the complexity hours and cost based on the complexity level and user settings.

    Parameters:
        userData (dict): The user data containing complexity settings and hourly rates.
        complexity_level (str): The determined complexity level for the task.

    Returns:
        dict: Estimated hours and cost for the task.
    """
    try:
        # Validate the required settings are available in userData
        if not userData or "settings" not in userData or "level_of_complexities" not in userData["settings"] or "hourly_rate" not in userData["settings"]:
            raise ValueError("Invalid userData structure. Missing settings, level_of_complexities, or hourly_rate.")

        # Retrieve complexity levels and hourly rates from user settings
        level_of_complexities = userData["settings"]["level_of_complexities"]
        hourly_rate = userData["settings"]["hourly_rate"]

        # Determine the complexity level
        complexity_level_str = determine_complexity_level(complexity_level)

        # Validate the complexity level exists in user settings
        if complexity_level_str not in level_of_complexities:
            raise ValueError(f"Complexity level '{complexity_level_str}' not found in level_of_complexities.")

        # Calculate estimated hours and cost
        estimated_hours = level_of_complexities[complexity_level_str]
        cost = estimated_hours * hourly_rate
        return {"cost": cost, "estimated_hours": estimated_hours}

    except ValueError as ve:
        print(f"Validation error: {ve}")
        raise ve
    except Exception as e:
        print("Error in calculating complexity hours", e)
        raise e
    
def update_existing_story(payload, user_id):
    """
    Service to update an existing story for a specific epic and stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, story_id, title, description, and acceptance_criteria.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with the updated story, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                epics_list = stakeholder_data["data"]

                # Find the epic in which the story exists
                for epic in epics_list:
                    if epic["id"] == payload["epic_id"]:
                        stories_list = epic.get("user_stories", [])

                        # Find the story to update by id
                        for story in stories_list:
                            if story["id"] == payload["story_id"]:
                                # Update the title, description, and acceptance_criteria of the selected story
                                story["title"] = payload["title"]
                                story["description"] = payload["description"]
                                story["acceptance_criteria"] = payload["acceptance_criteria"]

                                # Update the proposal in the database with the updated story
                                db.proposals.find_one_and_update(
                                    {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                                    {
                                        "$set": {
                                            "epics.$[stakeholder].data.$[epic].user_stories": stories_list,
                                            "updated_at": datetime.datetime.utcnow()
                                        }
                                    },
                                    array_filters=[
                                        {"stakeholder.stakeholder": payload["stakeholder"]},
                                        {"epic.id": payload["epic_id"]}
                                    ]
                                )
                                # Optionally, remove tasks from the response if not needed
                                story_without_tasks = {k: v for k, v in story.items() if k != "tasks"}

                                # Return a success response with the updated story only
                                return make_response(
                                    status="success",
                                    message=STORY_UPDATED_SUCCESS,
                                    data=story_without_tasks,  # Only the story is returned in the response
                                    status_code=200
                                )

                        # If story_id is not found, return an error message
                        return make_response(
                            status="error",
                            message=STORY_NOT_FOUND,
                            data=None,
                            status_code=404
                        )

                # If epic_id is not found, return an error message
                return make_response(
                    status="error",
                    message=EPIC_NOT_FOUND,
                    data=None,
                    status_code=404
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in updating story: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def update_existing_task(payload, user_id):
    """
    Service to update an existing task for a specific story in an epic and stakeholder.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, story_id, task_id, description, complexity.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message and the updated task, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                epics_list = stakeholder_data["data"]

                # Find the specific epic in the stakeholder's data
                for epic in epics_list:
                    if epic["id"] == payload["epic_id"]:
                        stories_list = epic.get("user_stories", [])

                        # Find the specific story in the epic
                        for story in stories_list:
                            if story["id"] == payload["story_id"]:
                                tasks_list = story.get("tasks", [])

                                # Find the specific task to update
                                for task in tasks_list:
                                    if task["id"] == payload["task_id"]:
                                        # Update the task fields such as description, complexity, etc.
                                        task["description"] = payload.get("description", task["description"])
                                        task["complexity"] = payload.get("complexity", task["complexity"])
                                        # Fetch user data for complexity analysis
                                        user_data = db.users.find_one({"_id": ObjectId(user_id)})

                                        # Recalculate estimated hours and cost based on new complexity
                                        if "complexity" in payload:
                                            complexity_data = async_calculate_complexity_hours_for_add_task(user_data, task["complexity"])
                                            task["estimated_hours"] = complexity_data["estimated_hours"]
                                            task["cost"] = complexity_data["cost"]

                                        # Update the proposal in the database with the updated task
                                        db.proposals.find_one_and_update(
                                            {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                                            {
                                                "$set": {
                                                    "epics.$[stakeholder].data.$[epic].user_stories.$[story].tasks": tasks_list,
                                                    "updated_at": datetime.datetime.utcnow()
                                                }
                                            },
                                            array_filters=[
                                                {"stakeholder.stakeholder": payload["stakeholder"]},
                                                {"epic.id": payload["epic_id"]},
                                                {"story.id": payload["story_id"]}
                                            ]
                                        )

                                        # Return a success response with the updated task details
                                        return make_response(
                                            status="success",
                                            message=TASK_UPDATED_SUCCESS,
                                            data={"task": task},
                                            status_code=200
                                        )

                                # If task_id is not found, return an error message
                                return make_response(
                                    status="error",
                                    message=TASK_NOT_FOUND,
                                    data=None,
                                    status_code=404
                                )

                        # If story_id is not found, return an error message
                        return make_response(
                            status="error",
                            message=STORY_NOT_FOUND,
                            data=None,
                            status_code=404
                        )

                # If epic_id is not found, return an error message
                return make_response(
                    status="error",
                    message=EPIC_NOT_FOUND,
                    data=None,
                    status_code=404
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in updating task: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def delete_existing_story(payload, user_id):
    """
    Service to delete an existing story from a specific epic within a stakeholder in a proposal.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, and story_id.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == payload["epic_id"]:
                        # Get the current list of stories
                        stories_list = epic.get("user_stories", [])

                        # Find and remove the story by story_id
                        story_to_delete = next((story for story in stories_list if story["id"] == payload["story_id"]), None)

                        if story_to_delete:
                            stories_list.remove(story_to_delete)

                            # Update the epic in the proposal with the updated stories list
                            db.proposals.find_one_and_update(
                                {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                                {
                                    "$set": {
                                        "epics.$[stakeholder].data.$[epic].user_stories": stories_list,
                                        "updated_at": datetime.datetime.utcnow()
                                    }
                                },
                                array_filters=[
                                    {"stakeholder.stakeholder": payload["stakeholder"]},
                                    {"epic.id": payload["epic_id"]}
                                ]
                            )

                            # Return a success response
                            return make_response(
                                status="success",
                                message=STORY_DELETED_SUCCESS,
                                data=None,
                                status_code=200
                            )

        # If the story or epic is not found, return an error message
        return make_response(
            status="error",
            message=STORY_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in deleting story: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def delete_existing_task(payload, user_id):
    """
    Service to delete an existing task from a specific story within a stakeholder's epic.

    Parameters:
        payload (dict): Contains proposal_id, stakeholder, epic_id, story_id, and task_id.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with a success message or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == payload["stakeholder"]:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == payload["epic_id"]:
                        # Find the specific story in the epic's user_stories
                        for story in epic.get("user_stories", []):
                            if story["id"] == payload["story_id"]:
                                # Find and remove the task by task_id
                                task_to_delete = next((task for task in story.get("tasks", []) if task["id"] == payload["task_id"]), None)

                                if task_to_delete:
                                    # Remove the task
                                    story["tasks"].remove(task_to_delete)

                                    # Update the proposal in the database
                                    db.proposals.find_one_and_update(
                                        {"_id": ObjectId(payload["proposal_id"]), "user": ObjectId(user_id)},
                                        {
                                            "$set": {
                                                "epics.$[stakeholder].data.$[epic].user_stories.$[story].tasks": story["tasks"],
                                                "updated_at": datetime.datetime.utcnow()
                                            }
                                        },
                                        array_filters=[
                                            {"stakeholder.stakeholder": payload["stakeholder"]},
                                            {"epic.id": payload["epic_id"]},
                                            {"story.id": payload["story_id"]}
                                        ]
                                    )

                                    # Return a success response
                                    return make_response(
                                        status="success",
                                        message=TASK_DELETED_SUCCESS,
                                        data=None,
                                        status_code=200
                                    )

        # If the task or story is not found, return an error message
        return make_response(
            status="error",
            message=TASK_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in deleting task: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def fetch_stories_by_epic_and_stakeholder(proposal_id, stakeholder, epic_id, user_id):
    """
    Service to fetch all stories for a specific epic and stakeholder in a proposal,
    excluding tasks from the response.

    Parameters:
        proposal_id (str): The ID of the proposal.
        stakeholder (str): The name of the stakeholder.
        epic_id (str): The ID of the epic.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with the list of stories excluding tasks or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == stakeholder:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == epic_id:
                        # Get the list of stories and exclude tasks
                        stories_without_tasks = []
                        for story in epic.get("user_stories", []):
                            story_copy = story.copy()  # Create a copy to modify
                            if "tasks" in story_copy:
                                del story_copy["tasks"]  # Remove tasks from the story
                            stories_without_tasks.append(story_copy)

                        # Return a success response with the modified list of stories
                        return make_response(
                            status="success",
                            message=STORIES_FETCHED_SUCCESS,
                            data={"stories": stories_without_tasks},
                            status_code=200
                        )

        # If the stakeholder or epic is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_OR_EPIC_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in fetching stories: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def fetch_tasks_by_story(proposal_id, stakeholder, epic_id, story_id, user_id):
    """
    Service to fetch tasks for a specific story in an epic within a stakeholder in a proposal.

    Parameters:
        proposal_id (str): The ID of the proposal.
        stakeholder (str): The name of the stakeholder.
        epic_id (str): The ID of the epic.
        story_id (str): The ID of the story.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A response with the list of tasks or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})

        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == stakeholder:
                # Find the specific epic in the stakeholder's data
                for epic in stakeholder_data["data"]:
                    if epic["id"] == epic_id:
                        # Find the specific story in the epic's user stories
                        for story in epic.get("user_stories", []):
                            if story["id"] == story_id:
                                # Get the tasks for the story
                                tasks = story.get("tasks", [])
                                
                                # Return a success response with the list of tasks
                                return make_response(
                                    status="success",
                                    message=TASKS_FETCHED_SUCCESS,
                                    data={"tasks": tasks},
                                    status_code=200
                                )

        # If the stakeholder, epic, or story is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_EPIC_OR_STORY_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in fetching tasks: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def fetch_epics(proposal_id, stakeholder, user_id):
    """
    Fetch epics for a specific stakeholder in a proposal, excluding user stories.

    Parameters:
        proposal_id (str): The ID of the proposal.
        stakeholder (str): The name of the stakeholder.
        user_id (str): The ID of the user making the request.

    Returns:
        dict: A response with a list of epics excluding user stories, or an error message.
    """
    try:
        # Fetch the proposal from the database by proposal_id and user_id
        proposal = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        
        # Check if the proposal exists
        if not proposal:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )
        
        # Find the stakeholder's epics list in the proposal
        for stakeholder_data in proposal["epics"]:
            if stakeholder_data["stakeholder"] == stakeholder:
                epics_list = stakeholder_data["data"]
                
                # Remove user_stories from each epic in the list
                filtered_epics = []
                for epic in epics_list:
                    filtered_epic = {
                        "epic_id": epic.get("epic_id"),
                        "title": epic.get("title"),
                        "description": epic.get("description"),
                        "id": epic.get("id")
                    }
                    filtered_epics.append(filtered_epic)

                # Return a success response with the filtered epics
                return make_response(
                    status="success",
                    message=EPICS_FETCHED_SUCCESS,
                    data={"epics": filtered_epics},
                    status_code=200
                )

        # If the stakeholder is not found, return an error message
        return make_response(
            status="error",
            message=STAKEHOLDER_NOT_FOUND,
            data=None,
            status_code=404
        )

    except Exception as e:
        print(f"Error in fetching epics: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def delete_proposal(proposal_id, user_id):
    """
    Service to delete a proposal and its related conversation based on proposal ID and user ID.

    Parameters:
        proposal_id (str): The ID of the proposal to be deleted.
        user_id (str): The ID of the user requesting the deletion.

    Returns:
        dict: A response indicating whether the deletion was successful or not.
    """
    try:
        # Convert the proposal_id to ObjectId
        proposal_object_id = ObjectId(proposal_id)

        # Attempt to find and delete the proposal associated with the user
        proposal_result = db.proposals.find_one_and_delete({"_id": proposal_object_id, "user": user_id})

        # If no proposal was found, return a 404 error
        if proposal_result is None:
            return make_response(
                status="error",
                message=PROPOSAL_NOT_FOUND,
                data=None,
                status_code=404
            )

        # Attempt to find and delete the conversation related to the proposal and user
        conversation_result = db.conversations.find_one_and_delete(
            {"proposal_id": proposal_object_id, "user": user_id}
        )

        # Log if no conversation was found for the proposal
        if conversation_result is None:
            print(f"No conversation found for proposal ID {proposal_id}")

        # Return a success message after successful deletion
        return make_response(
            status="success",
            message="Proposal and related conversation deleted successfully.",
            data=None,
            status_code=200
        )

    except Exception as e:
        # Handle unexpected exceptions and return a 500 Internal Server Error response
        print(f"Error in deleting proposal and conversation: {e}")
        return make_response(
            status="error",
            message=INTERNAL_SERVER_ERROR,
            data=None,
            status_code=500
        )

def generate_proposal_details(proposal_id, user_id):
    """Generate Proposal Details for the user based on the proposal id and user id"""
    try:
        data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": user_id})
        if not data:
            raise ValueError("No data found for the given proposal_id and user_id")

        if "epics" not in data or "project_requirement" not in data:
            raise ValueError("Required keys ('epics' and 'project_requirement') not found in the data")

        # Fetch the profile data
        profile_data = db.individual_profiles.find_one({"user_id": ObjectId(user_id)})
        
        # If profile_data is not available, set related info to blank
        if not profile_data:
            profile_data = {
                'company_name': '',
                'name': '',
                'address': '',
                'email': '',
                'phone_number': '',
            }
        
        project_breakdown = data["epics"]
        project_requirement = data["project_requirement"]

        project_details = Proposal_Creation(project_breakdown, project_requirement, user_id, proposal_id)
        global_deliverables = project_details["Deliverables"]

        project_milestone = Proposal_Milestones(
            project_breakdown, project_requirement, global_deliverables, user_id, proposal_id
        )
        project_details["Milestones"] = project_milestone["Milestones"]

        project_risks = Proposal_Risks(project_breakdown, project_requirement, user_id, proposal_id)
        project_details["Risks"] = project_risks["Risks"]

        budget = calculate_costs_and_hours(project_breakdown)
        project_details["Budget"] = budget

        project_details["Date"] = {"CurrentDate": datetime.datetime.today().strftime("%Y-%m-%d")}

        project_details["ProposalID"] = data["sr_no"]
        logo_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../utils/logos/Logo.png'))

        if not os.path.exists(logo_path):
            return "Logo file not found", 404

        with open(logo_path, "rb") as image_file:
            encoded_logo = base64.b64encode(image_file.read()).decode('utf-8')

        # Populate ClientInformation with profile data
        project_details["ClientInformation"] = {
            "CompanyName": profile_data['company_name'],
            "ContactPerson": profile_data['name'],
            "Address": profile_data['address'],
            "Email": profile_data['email'],
            "Phone": profile_data['phone_number'],
            "logo": encoded_logo
        }

        # Static User Information
        project_details["UserInformation"] = {
            "CompanyName": "Myte Group Inc.",
            "Address": "7501 M.B Jodoin, Anjou, ",
            "Email": "ahmed.mekallach@mytegroup.com",
            "Phone": "5148049207",
            "Website": "www.mytegroup.com",
        }

        # Our Qualifications section
        project_details["OurQualifications"] = {
            "CompanyProfile": "At Myte, we specialize in crafting custom AI automation solutions that streamline workflows and enhance operational efficiency for businesses across various industries. Our expertise lies in developing AI-powered websites and digital work environments tailored to the unique needs of our clients. By integrating artificial intelligence into core business processes, we deliver systems that not only reduce manual effort but also significantly increase accuracy and decision-making speed. Our commitment to innovation and excellence empowers businesses to achieve sustainable growth and maintain a competitive edge in their respective markets.",
            "RelevantExperience": "Myte Group has extensively optimized various internal operations through customized AI-driven solutions, particularly in sales, marketing, outreach, estimation, planning, and research. This expertise reflects our ability to enhance process efficiencies and accuracy, thereby improving time management and resource allocation across departments. Our strategic implementations of AI have significantly streamlined these key business functions, demonstrating our capability to elevate organizational performance and drive substantial growth through innovation in AI technology.",
            "TeamExpertise": "Under the leadership of Ahmed Mekallach, Myte Group boasts a profound expertise uniquely blending business acumen with technical prowess. With over eight years of experience in sales, estimating, and project management, Ahmed's comprehensive understanding of business processes and keen coding skills enable him to tackle challenges efficiently and innovatively. His approach not only ensures operational excellence but also drives the development of AI-driven solutions that are both sophisticated and practical, perfectly suited to meet the intricate demands of modern businesses. This unique combination positions Myte Group as a leader in AI systems design and development, capable of delivering high-quality, customized solutions.",
        }

        return project_details
    except Exception as e:
        print("Error in generating proposal details", e)
        raise e

def Proposal_Creation(project_breakdown, project_requirement,user_id, praposal_id):
    try:
        system_context = proposal_system_context
        assistant_context = proposal_assistant_context.replace(
            "{proposal_json_format}", proposal_json_format
        )
        #input_message = proposal_initial_prompt.replace(
            #"{project_breakdown}", str(project_breakdown)
        #)
        input_message = proposal_initial_prompt.replace(
            "{project_requirements}", str(project_requirement)
        )
        project_details = json.loads(
            generate_text_json(
                system_context,
                assistant_context,
                input_message,
                user_id, 
                praposal_id
            )
        )

        return project_details
    except Exception as e:
        print("Error in generating proposal details", e)
        return 0


def Proposal_Milestones(project_breakdown, project_requirement, global_deliverables,user_id, proposal_id):
    try:
        system_context = milestone_system_context
        assistant_context = milestone_assistant_context.replace(
            "{milestone_json_format}", milestone_json_format
        )
        #input_message = milestone_initial_prompt.replace(
            #"{project_breakdown}", str(project_breakdown)
        #)
        input_message = milestone_initial_prompt.replace(
            "{project_requirements}", str(project_requirement)
        )
        input_message = input_message.replace(
            "{global_deliverables}", str(global_deliverables)
        )

        milestone = json.loads(
            generate_text_json(
                system_context,
                assistant_context,
                input_message,
                user_id, 
                proposal_id
            )
        )

        return milestone
    except Exception as e:
        print("Error in generating proposal milestones", e)
        return 0


def Proposal_Risks(project_breakdown, project_requirement,user_id, proposal_id):
    """Generate Proposal Risks based on the project breakdown and project requirements"""
    try:
        system_context = risks_system_context
        assistant_context = risks_assistant_context.replace(
            "{risks_json_format}", risks_json_format
        )
        #input_message = risks_initial_prompt.replace(
            #"{project_breakdown}", str(project_breakdown)
        #)
        input_message = risks_initial_prompt.replace(
            "{project_requirements}", str(project_requirement)
        )
        risks = json.loads(
            generate_text_json(
                system_context,
                assistant_context,
                input_message,
                user_id, 
                proposal_id
            )
        )

        return risks
    except Exception as e:
        print("Error in generating proposal risks", e)
        return 0

def calculate_costs_and_hours(project_breakdown):
    total_hours = 0
    total_cost = 0
    cost_breakdown = []

    # Iterate over each stakeholder and their epics
    for epics in project_breakdown:
        stakeholder_details = {"Stakeholder": epics["stakeholder"], "Details": []}

        # Iterate over each epic and its stories
        for epic_data in epics["data"]:
            epic_hours = 0
            epic_cost = 0
            epic_description = ""

            # Process each story in the epic
            for story in epic_data["user_stories"]:
                story_hours = sum(task["estimated_hours"] for task in story["tasks"])
                story_cost = sum(task["cost"] for task in story["tasks"])
                if (
                    not epic_description
                ):  # Assume all stories share the same epic description
                    epic_description = epic_data["description"]
                epic_hours += float(story_hours)
                epic_cost += float(story_cost)

            stakeholder_details["Details"].append(
                {
                    "Item": epic_data["user_stories"][0][
                        "title"
                    ],  # Assuming the first story's title represents the epic
                    "TotalHours": epic_hours,
                    "Cost": epic_cost,
                    "Description": epic_description,
                }
            )

            total_hours += epic_hours
            total_cost += epic_cost

        cost_breakdown.append(stakeholder_details)

    return {
        "TotalEstimatedHours": total_hours,
        "TotalEstimatedCost": total_cost,
        "CostBreakdown": cost_breakdown,
    }

def generateDetailReport(proposal_id, user_id):
    try:
        # Generate JSON data for proposal details
        json_data = generate_proposal_details(proposal_id, user_id)
        # print("🚀 ~ json_data:", json_data)
        
        pdf_content = design_pdf(json_data)  # Generate PDF from JSON data

        # Fetch proposal and user data
        data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": user_id})
        user = db.users.find_one({"_id": ObjectId(user_id)})

        # Prepare rows for the proposal detail report
        rows = []
        previous_story_id = None  # Track the previous story ID
        for epic in data["epics"]:
            stakeholder = epic["stakeholder"]
            for epics_data in epic["data"]:
                epic_id = epics_data["epic_id"]
                epic_title = epics_data["title"]
                epic_description = epics_data["description"]
                for story in epics_data["user_stories"]:
                    acceptance_criteria = "\n".join(story['acceptance_criteria'])
                    if story['story_id'] != previous_story_id:  # Check if it's a new story
                        story_id = story['story_id']
                        story_title = story['title']
                        story_description = story['description']
                        task_counter = 0  # Reset task counter for each new story
                    # Sort the tasks by 'task_id'
                    sorted_tasks = sorted(story['tasks'], key=lambda task: task['task_id'])
                    for task in sorted_tasks:
                        task_counter += 1
                        task_id = task['task_id']  # Use task ID from data
                        row = {
                            'Stakeholder': stakeholder if task_counter == 1 else '',
                            'EpicID': epic_id if task_counter == 1 else '',
                            'EPIC TITLE': epic_title if task_counter == 1 else '',
                            'EPIC DESCRIPTION': epic_description if task_counter == 1 else '',
                            'Story ID': story_id if task_counter == 1 else '',
                            'Story Title': story_title if task_counter == 1 else '',
                            'Story Description': story_description if task_counter == 1 else '',
                            'Story Acceptance Criteria': acceptance_criteria if task_counter == 1 else '',
                            'Task ID': task_id,
                            'Task Description': task['description'],
                            'Estimated Hours': task['estimated_hours'],
                            'Cost': task['cost']
                        }
                        rows.append(row)
                    previous_story_id = story['story_id']  # Update the previous story ID

        # Convert rows to DataFrame
        df = pd.DataFrame(rows)

        # Create a new Excel file with the DataFrame
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proposal Details"

        # Append DataFrame rows to Excel sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Custom Style for readability
        header_style = NamedStyle(name="header_style", font=Font(bold=True, color="FFFFFF"), fill=PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"))
        ws.row_dimensions[1].height = 20
        for cell in ws[1]:
            cell.style = header_style

        # Apply word wrap to specific columns
        wrap_alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column_letter in ['C', 'D', 'F', 'G', 'H', 'J']:
                    cell.alignment = wrap_alignment


        def apply_merge_style(ws, column_indices):
            for idx in column_indices:
                current_cell = ws.cell(row=2, column=idx)
                start_row = 2

                for row in range(3, ws.max_row + 1):
                    cell = ws.cell(row=row, column=idx)
                    
                    # Check if the current cell value is different from the previous cell or if it's the last row
                    if (cell.value != current_cell.value and cell.value != "") or row == ws.max_row:
                        end_row = row - 1 if cell.value != current_cell.value else row
                        
                        # Merge cells only if there are multiple rows to merge
                        if start_row != end_row:
                            ws.merge_cells(start_row=start_row, start_column=idx, end_row=end_row, end_column=idx)
                        
                        # Apply alignment for the merged cells
                        for r in range(start_row, end_row + 1):
                            ws.cell(row=r, column=idx).alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

                        # Update the current cell and start row
                        current_cell = cell
                        start_row = row
                
                # Handle the case where the last rows have the same value or are blank
                if start_row != ws.max_row:
                    ws.merge_cells(start_row=start_row, start_column=idx, end_row=ws.max_row, end_column=idx)
                    for r in range(start_row, ws.max_row + 1):
                        ws.cell(row=r, column=idx).alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)


        # Apply merging to all relevant columns including the new Stakeholder column
        apply_merge_style(ws, [1, 2, 3, 4, 5, 6, 7, 8])

        # Adjust column widths
        column_widths = {
            'A': 20,
            'B': 10,
            'C': 30,
            'D': 50,
            'E': 10,
            'F': 30,
            'G': 50,
            'H': 50,
            'I': 10,
            'J': 40,
            'K': 15,
            'L': 15
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add a new worksheet for the conversation data
        ws_conversation = wb.create_sheet(title="Conversation")

        # Fetch conversation data
        conversation_data = db.conversation.find_one({"user_id": ObjectId(user_id), "proposal_id": ObjectId(proposal_id)})
        
        # Prepare rows for the conversation report
        conversation_rows = []
        if conversation_data:
            for convo in conversation_data["data"]:
                for key, value in convo.items():
                    conversation_rows.append({'Speaker': key, 'Message': value})

        # Convert conversation rows to DataFrame
        df_conversation = pd.DataFrame(conversation_rows)

        # Append DataFrame rows to the Conversation sheet
        for r in dataframe_to_rows(df_conversation, index=False, header=True):
            ws_conversation.append(r)

        # Apply similar styling to the conversation sheet
        ws_conversation.row_dimensions[1].height = 20
        for cell in ws_conversation[1]:
            cell.style = header_style

        # Adjust column widths in the conversation sheet
        ws_conversation.column_dimensions['A'].width = 20
        ws_conversation.column_dimensions['B'].width = 80

        # Apply word wrap to column B in the conversation sheet
        for row in ws_conversation.iter_rows(min_row=2, max_row=ws_conversation.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = wrap_alignment
        #file name
        proposal_data = json_data
        client_company_name = sanitize_filename(proposal_data['ClientInformation']['CompanyName'])
        project_title = sanitize_filename(proposal_data['ProjectTitle'])
        proposal_date = proposal_data['Date']['CurrentDate']
        sanitized_proposal_id = str(proposal_data['ProposalID']).replace('-', '')
        folder_name = f"S{sanitized_proposal_id} - {project_title} - {client_company_name} - {proposal_date}"
        file_name = f"S{sanitized_proposal_id} - {project_title} - Proposal - {proposal_date}"
        report_directory = os.path.join(os.getcwd(), 'report')

        # Save the Excel file
        # sanitized_proposal_id = proposal_id.replace('-', '')
        file_name = f"S{file_name} - Proposal Detail Report.xlsx"
        wb.save(file_name)

        body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background-color: #0d0d0d;
                    color: #ffffff;
                    margin: 0;
                    padding: 0;
                    text-align: center;
                }}
                .container {{
                    max-width: 600px;
                    margin: 40px auto;
                    padding: 20px;
                    background-color: white;
                    border-radius: 10px;
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.5);
                    text-align: center;
                    border: 1px solid;

                }}
                .logo {{
                    margin-bottom: 20px;
                    background-color:#192655
                }}
                .logo img {{
                    width: 200px;
                    height: 90px;
                    display: block;
                    margin: 0 auto;
                }}
                .greeting {{
                    font-size: 18px;
                    margin-bottom: 20px;
                    line-height: 1.5;
                    color: black;
                }}
                .content {{
                    font-size: 16px;
                    line-height: 1.5;
                    color: black;
                    margin-bottom: 20px;
                }}
                .footer {{
                    margin-top: 40px;
                    font-size: 12px;
                    color: black;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="logo">
                    <img src="https://myte-social-subscription.s3.ca-central-1.amazonaws.com/asset/myte-logo.png" width="161.87px" height="86.6px"  alt="Myte Cody Logo">
                </div>
                <div class="greeting">
                    <p><b>Hello,</b></p>
                    <p><b>We are excited to share the proposal details generated by Myte Cody - our AI Software Planner and Estimator.</b></p>
                </div>
                <div class="content">
                    <p><b>Attached, you will find the proposal that outlines the detailed roadmap and estimations for your project.</b></p>
                    <p><b>Please review the attached documents carefully, and feel free to reach out if you have any questions or need further clarifications.</b></p>
                </div>
                <div class="footer">
                    <p>If you did not request this proposal, please ignore this email.</p>
                </div>
            </div>
        </body>
        </html>
        """
        # Email configuration
        sender_email = os.environ.get("EMAIL")
        recipient_email = user['email']
        subject = "Myte Cody-" + json_data['ProjectTitle'] + " - " + datetime.datetime.today().strftime("%Y-%m-%d")
        # body = "Please find the proposal details attached."

        # Create a multipart message and set headers
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient_email
        message["Subject"] = subject

        # # Add body to email
        # message.attach(MIMEText(body, "plain"))

        # Attach HTML body
        message.attach(MIMEText(body, "html"))

        # Attach Excel file
        with open(file_name, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {file_name}",
        )
        message.attach(part)

        pdf_filename = f"S{file_name}.pdf"
        part = MIMEBase("application", "octet-stream")
        part.set_payload(pdf_content)
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {pdf_filename}",
        )
        message.attach(part)


        # Convert message to string
        text = message.as_string()

        # Log in to email server
        server = smtplib.SMTP(os.environ.get("SMTP_SERVER"), 587)
        server.starttls()
        server.login(sender_email, os.environ.get("EMAIL_PASSWORD"))

        # Send email
        server.sendmail(sender_email, recipient_email, text)
        server.quit()

        db.proposals.find_one_and_update(
            {"user": ObjectId(user_id), "_id": ObjectId(proposal_id)},
            {
                "$set": {
                    "status": PROPOSAL_STATUS["completed"],
                    "step": 6,
                }
            },
        )
        return {"message": "Success"}
    except Exception as e:
        print("Error in generating report", e)
        raise e
def fetch_proposals(user_id, page=1, status=None):
    """
    Service to fetch paginated proposals for a specific user with optional status filtering.

    Parameters:
        user_id (str): The ID of the user making the request.
        page (int): The page number for pagination (default is 1).
        status (str): The optional status filter for proposals ('in-progress', 'completed').

    Returns:
        dict: A JSON response with paginated proposal data, tokens used, or an error message.
    """
    try:
        # Set up pagination variables
        per_page = 10
        skip = (page - 1) * per_page

        # Create a filter based on the status if provided
        query_filter = {"user": ObjectId(user_id)}
        if status and status != "all":  # Allow 'all' to bypass status filtering
            query_filter["status"] = status

        # Define the projection to limit fields returned
        projection_fields = {
            "_id": 1,
            "business_vertical": 1,
            "last_step": 1,
            "status": 1,
            "user": 1,
            "step": 1,
            "sr_no": 1,
            "title": 1,
            "description": 1
        }

        # Fetch proposals from the database with pagination and projection
        proposals = list(db.proposals.find(query_filter, projection_fields).skip(skip).limit(per_page))

        if not proposals:
            return make_response(
                status="error",
                message="No proposals found",
                data=None,
                status_code=404
            )

        # Convert ObjectId to string in each proposal
        for proposal in proposals:
            proposal["_id"] = str(proposal["_id"])
            proposal["user"] = str(proposal["user"])

            # Add tokens_used from the proposal_usage collection
            usage = db.proposal_usage.find_one(
                {"proposal_id": ObjectId(proposal["_id"]), "user_id": ObjectId(user_id)},
                {"tokens_used": 1, "_id": 0}
            )
            proposal["tokens_used"] = usage.get("tokens_used", 0) if usage else 0

        # Return the proposals in the response with pagination metadata
        total_proposals = db.proposals.count_documents(query_filter)
        response_data = {
            "proposals": proposals,
            "total": total_proposals,
            "page": page,
            "per_page": per_page,
            "total_pages": (total_proposals + per_page - 1) // per_page
        }

        return make_response(
            status="success",
            message="Proposals fetched successfully",
            data=response_data,
            status_code=200
        )

    except Exception as e:
        print(f"Error in fetching proposals: {e}")
        return make_response(
            status="error",
            message="Internal Server Error",
            data=None,
            status_code=500
        )

def proposal_basic_info(proposal_id, user_id):
    """
    Service function to generate basic proposal details for the given proposal_id and user_id.

    Parameters:
        proposal_id (str): The ID of the proposal.
        user_id (str): The ID of the user making the request.

    Returns:
        (dict): A JSON response with proposal details, budget information, and token usage, or an error message.
    """
    try:
        # Fetch the proposal based on proposal_id and user_id
        data = db.proposals.find_one({"_id": ObjectId(proposal_id), "user": ObjectId(user_id)})
        
        if not data:
            return make_response(
                status="error",
                message="Proposal not found for the given proposal_id and user_id",
                data=None,
                status_code=404
            )

        # Ensure the required keys exist in the data
        if "epics" not in data or "project_requirement" not in data:
            return make_response(
                status="error",
                message="Missing required fields 'epics' or 'project_requirement' in the proposal data",
                data=None,
                status_code=400
            )
        
        # Extract the project breakdown and requirement
        project_breakdown = data["epics"]
        project_requirement = data["project_requirement"]

        # Call helper function to create detailed proposal information
        project_details = Proposal_Creation(project_breakdown, project_requirement, user_id, proposal_id)

        # Calculate budget (costs and hours) and add it to the proposal details
        budget = calculate_total_costs_and_hours(project_breakdown)
        project_details["Budget"] = budget
        
        # Fetch the tokens used from the proposal_usage collection
        tokens_used_data = db.proposal_usage.find_one(
            {"proposal_id": ObjectId(proposal_id), "user_id": ObjectId(user_id)}
        )
        project_details["token_used"] = tokens_used_data.get("tokens_used", 0) if tokens_used_data else 0

        # Return success response with proposal details
        return make_response(
            status="success",
            message="Proposal report generated successfully",
            data=project_details,
            status_code=200
        )
    
    except ValueError as ve:
        # Handle known errors (like missing data)
        print(f"ValueError: {ve}")
        return make_response(
            status="error",
            message=str(ve),
            data=None,
            status_code=400
        )
    
    except Exception as e:
        # Handle unexpected errors
        print(f"Error in generating proposal details: {e}")
        return make_response(
            status="error",
            message="Internal Server Error",
            data=None,
            status_code=500
        )

def calculate_total_costs_and_hours(project_breakdown):
    """
    Calculate the total estimated hours and total estimated cost for the given project breakdown.

    Parameters:
        project_breakdown (list): A list of epics, each containing user stories with tasks.

    Returns:
        dict: A dictionary containing the total estimated hours and total estimated cost.
    """
    total_hours = 0
    total_cost = 0

    # Iterate over each epic in the project breakdown
    for epics in project_breakdown:
        # Iterate over each epic's stories
        for epic_data in epics["data"]:
            # Process each story in the epic
            for story in epic_data["user_stories"]:
                # Calculate total hours and cost for the story's tasks
                story_hours = sum(float(task["estimated_hours"]) for task in story["tasks"])
                story_cost = sum(float(task["cost"]) for task in story["tasks"])
                
                # Accumulate the total hours and cost
                total_hours += story_hours
                total_cost += story_cost

    # Return the total estimated hours and cost
    return {
        "TotalEstimatedHours": total_hours,
        "TotalEstimatedCost": total_cost,
    }
